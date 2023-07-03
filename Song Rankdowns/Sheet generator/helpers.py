import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side
import inflect
p = inflect.engine()


def get_color_ranges(length, colors):
    colors_copy = colors.copy()
    colors_copy.reverse()
    color_list = []
    for i in range(length):
        color_list.append(colors_copy[i % len(colors_copy)])

    final_list = []
    jumpy_index = 0
    tracker = 0
    for i in range(length):
        if jumpy_index >= length:
            tracker += 1
            jumpy_index = tracker 

        final_list.append(color_list[jumpy_index])

        jumpy_index += len(colors_copy)

    final_list.reverse()
    return final_list


def get_color_ranges_with_specials(length, colors, specials):
    return get_color_ranges(length, colors) + specials


def unicolor_column(sheet, column, how_long, width, color):
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    
    for i in range(how_long):
        sheet.cell(row=i+1, column=column).value = ""
        sheet.cell(row=i+1, column=column).fill = fill
        
    sheet.column_dimensions[sheet.cell(row=1, column=column).column_letter].width = width


def unicolor_column2(sheet, column, start, end, width, color):
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    
    for i in range(start, end):
        sheet.cell(row=i+1, column=column).value = "hi"
        sheet.cell(row=i+1, column=column).fill = fill
        
    sheet.column_dimensions[sheet.cell(row=1, column=column).column_letter].width = width


def apply_color(cell, color):
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.fill = fill



def decision_space(sheet, r, c, count, col_prim, col_sec, users) -> int:
    """
    Returns the row number of the next empty row
    """
    fill_prim = PatternFill(start_color=col_prim, end_color=col_prim, fill_type="solid")
    fill_sec = PatternFill(start_color=col_sec, end_color=col_sec, fill_type="solid")

    sheet.cell(row=r, column=c).value = f"{p.ordinal(count)} Place Vote:"
    sheet.cell(row=r, column=c).fill = fill_prim
    sheet.cell(row=r, column=c+1).value = "Explain Why You're Voting This Song:"
    sheet.cell(row=r, column=c+1).fill = fill_prim

    for k in range(len(users)):
        sheet.cell(row=r+1+k, column=c).value = f"{users[k]}'s Vote: "
        sheet.cell(row=r+1+k, column=c).fill = fill_sec
        sheet.cell(row=r+1+k, column=c+1).value = "opinion"
        sheet.cell(row=r+1+k, column=c+1).fill = fill_sec

    elim_row = r+1 + len(users)
    sheet.cell(row=elim_row, column=c).value = "Eliminated Song: "
    empty_row = elim_row+1
    sheet.cell(row=empty_row, column=c).value = "EMPTY"

    free_row = empty_row+1
    return free_row

