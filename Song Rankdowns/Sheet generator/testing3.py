from openpyxl import Workbook
from openpyxl.styles import PatternFill, Color

def apply_color(cell, color):
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.fill = fill

# Modify the transparency of a color
def modify_transparency(color, transparency):
    # Extract RGB values
    red = color.rgb.r
    green = color.rgb.g
    blue = color.rgb.b

    # Create a new Color object with modified transparency
    modified_color = Color(rgb=f"00{red:02X}{green:02X}{blue:02X}", tint=transparency)

    return modified_color

# Create a new workbook
workbook = Workbook()

# Access the active sheet
sheet = workbook.active

# Example usage
cell = sheet.cell(row=1, column=1)

# Original color
original_color = "FFFF0000"  # Red color (hex format)

# Modify transparency
transparency = 0.5  # Adjust the transparency level (0.0 - fully transparent, 1.0 - fully opaque)

# Convert original color to Color object
color = Color(rgb=original_color)

# Modify transparency of the color
modified_color = modify_transparency(color, transparency)

# Apply modified color to the cell
apply_color(cell, modified_color)

# Save the workbook
workbook.save("example.xlsx")
