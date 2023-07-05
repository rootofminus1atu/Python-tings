from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Color
import inflect
p = inflect.engine()
from general_helpers import *
from config import *

def apply_formatting(sheet, r, c, color=None, bold=False, italic=False, underline=None):
    cell = sheet.cell(row=r, column=c)

    font = Font(bold=bold, italic=italic, underline=underline)
    cell.font = font

    if color is None:
        return
    
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.fill = fill

    # mix the color with the default gridline color
    # that way it looks like a google sheets grid
    # just an idea


def change_width(sheet, c, width):
    sheet.column_dimensions[sheet.cell(row=1, column=c).column_letter].width = width

def song_list(sheet, r, c, songs):
    change_width(sheet, c, size_guide["big"])

    sheet.cell(row=r, column=c).value = "List of Songs:"
    apply_formatting(sheet, r, c, color_guide["header_blue"]["primary"], bold=True)

    for i, song in enumerate(songs):
        sheet.cell(row=r+1+i, column=c).value = song
        apply_formatting(sheet, r+1+i, c, color_guide["yellow"]["secondary"])

def scoreboard(sheet, r, c, color_ranges):
    first = c
    second = c + 1
    change_width(sheet, first, size_guide["small"])
    change_width(sheet, second, size_guide["big"])

    sheet.cell(row=r, column=first).value = "Place:"
    apply_formatting(sheet, r, first, color_guide["header_blue"]["primary"], bold=True)
    sheet.cell(row=r, column=second).value = "Score:"
    apply_formatting(sheet, r, second, color_guide["header_blue"]["primary"], bold=True)

    for i in range(len(songs)):
        sheet.cell(row=r+1+i, column=first).value = p.ordinal(len(songs) - i)
        apply_formatting(sheet, r+1+i, first, color_guide[color_ranges[i]]["primary"], bold=True)

    for i in range(len(songs)):
        sheet.cell(row=r+1+i, column=second).value = ""
        apply_formatting(sheet, r+1+i, second, color_guide[color_ranges[i]]["secondary"])

def decision_space(sheet, r, c, count, users, color_ranges):
    indx = len(color_ranges) - count

    color_prim = color_guide[color_ranges[indx]]["primary"]
    color_sec = color_guide[color_ranges[indx]]["secondary"]

    sheet.cell(row=r, column=c).value = f"{p.ordinal(count)} Place Vote: "
    apply_formatting(sheet, r, c, color_prim, bold=True, underline="single")
    sheet.cell(row=r, column=c+1).value = "Explain Why You're Voting This Song: "
    apply_formatting(sheet, r, c+1, color_prim, bold=True, underline="single")

    for i, user in enumerate(users):
        sheet.cell(row=r+1+i, column=c).value = f"{user}'s Vote: "
        apply_formatting(sheet, r+1+i, c, color_sec)
        sheet.cell(row=r+1+i, column=c+1).value = ""
        apply_formatting(sheet, r+1+i, c+1, color_sec)
    
    sheet.cell(row=r+1+len(users), column=c).value = "Eliminated Song: "
    apply_formatting(sheet, r+1+len(users), c, bold=True)

def nominated_songs(sheet, r, c, how_many_noms):
    sheet.cell(row=r, column=c).value = "Nominated Songs:"
    apply_formatting(sheet, r, c, color_guide["header_blue"]["primary"], bold=True)
    sheet.cell(row=r, column=c+1).value = "Nominated Song Link:"
    apply_formatting(sheet, r, c+1, color_guide["header_blue"]["primary"], bold=True)

    for i in range(how_many_noms):
        apply_formatting(sheet, r+1+i, c, color_guide["yellow"]["secondary"])
        apply_formatting(sheet, r+1+i, c+1, color_guide["yellow"]["secondary"])

def winner(sheet, r, c, color_ranges):
    sheet.cell(row=r, column=c).value = f"1st Place Winner: "
    apply_formatting(sheet, r, c, color_guide[color_ranges[-1]]["primary"], bold=True)
    apply_formatting(sheet, r, c+1, color_guide[color_ranges[-1]]["primary"], bold=True)

def black_border(sheet, c):
    change_width(sheet, c, size_guide["tiny"])

    how_long = 200
    for i in range(1, how_long):
        apply_formatting(sheet, i, c, color_guide["black"]["primary"])

def gray_border(sheet, c):
    change_width(sheet, c, size_guide["smaller"])

    how_long = 200
    for i in range(2, how_long):
        apply_formatting(sheet, i, c, color_guide["gray"]["primary"])
    apply_formatting(sheet, 1, c, color_guide["black"]["primary"])

def showtime(sheet, r, c, general_colors, special_colors, color_ranges, users, how_many_noms):
    nom_dec_gap = 2

    dec_start = 1 + how_many_noms + 1 + nom_dec_gap

    print(general_colors)
    nom_indexes = get_nomination_indexes(r, c, general_colors)
    dec_indexes = get_decision_indexes(dec_start, c, general_colors, special_colors, color_ranges, users)
    winner_index = dec_indexes.pop()
    print(nom_indexes)
    print(dec_indexes)
    print(winner_index)


    for row, col in nom_indexes:
        change_width(sheet, col, size_guide["big"])
        change_width(sheet, col+1, size_guide["huge"])
        nominated_songs(sheet, row, col, how_many_noms)

    counter = len(color_ranges)
    for row, col in dec_indexes:
        decision_space(sheet, row, col, counter, users, color_ranges)
        counter -= 1

    row, col = winner_index
    winner(sheet, row, col, color_ranges)

    for col in get_border_columns(nom_indexes):
        black_border(sheet, col)
    
