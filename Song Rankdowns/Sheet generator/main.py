import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from helpers import *
import inflect
p = inflect.engine()



# Create a new workbook
workbook = Workbook()

# Save the workbook to a file
workbook.save("output.xlsx")

# Load the workbook from the saved file
workbook = openpyxl.load_workbook("output.xlsx")

# Select the active sheet
sheet = workbook.active


"""# Add inputs to the sheet
sheet["A1"] = "Name"
sheet["B1"] = "Age"

sheet["A2"] = "Joooooooooooooooo"
sheet["B2"] = 25

sheet["A3"] = "Jane"
sheet["B3"] = 30


# Rescale column A to fit the content
sheet.column_dimensions['A'].width = 15

# Rescale column B to fit the content
sheet.column_dimensions['B'].width = 5"""








songs = ["hotline bling", "woohoo", "vacation", "bruja", "bruja", "bruja", "bruja", "bruja", "bruja", "bruja", "bruja", "bruja", "bruja", "bruja", "bruja", "bruja", "bruja", "bruja", "bruja", "anaconda", "MASS OF THE FERMENTING DERGS"]
users = ["joe", "jane", "sasha velour"]
nominations_num = 6

general_colors = ['red', 'orange', 'yellow','green']
top3_colors = ['bronze', 'silver', 'gold']
color_ranges = get_color_ranges(len(songs) - 3, general_colors) + top3_colors

max_black = 500  # will be replaced in the future

colors_guide = {
    "red": {
        "primary": "EA9999",
        "secondary": "F4CCCC"
    },
    "orange": {
        "primary": "F9CB9C",
        "secondary": "FCE5CD"
    },
    "yellow": {
        "primary": "FFE599",
        "secondary": "FFF2CC"
    },
    "green": {
        "primary": "B6D7A8",
        "secondary": "D9EAD3"
    },
    "bronze": {
        "primary": "B45F06",
        "secondary": "E69138"
    },
    "silver": {
        "primary": "B7B7B7",
        "secondary": "D9D9D9"
    },
    "gold": {
        "primary": "BF9000",
        "secondary": "F1C232"
    },
    "header_blue": {
        "primary": "4A86E8",
        "secondary": "4A86E8"
    },
    "black": {
        "primary": "000000",
        "secondary": "000000"
    },
    "gray": {
        "primary": "808080",
        "secondary": "808080"
    }
}





# coluumn 1 for gray squares
unicolor_column2(sheet, column=1, start=0, end=len(songs)+1, width=3, color="808080")
    
# column 2 for songs
for i, song in enumerate(songs):
    sheet.cell(row=i+2, column=2).value = song
    
# column 3 for black border
unicolor_column(sheet, 3, len(songs), 3, "000000")

def scoreboard(sheet, songs, color_ranges):
    # in D and E columns
    D = 4
    E = 5

    sheet.cell(row=1, column=D).value = "Place:"
    apply_color(sheet.cell(row=1, column=D), colors_guide["header_blue"]["primary"])
    sheet.cell(row=1, column=E).value = "Score:"
    apply_color(sheet.cell(row=1, column=E), colors_guide["header_blue"]["primary"])

    # D for place
    for i in range(len(songs)):
        sheet.cell(row=i+2, column=D).value = p.ordinal(len(songs) - i)
        apply_color(sheet.cell(row=i+2, column=D), colors_guide[color_ranges[i]]["primary"])

    # E for songs
    for i in range(len(songs)):
        sheet.cell(row=i+2, column=E).value = ""
        apply_color(sheet.cell(row=i+2, column=E), colors_guide[color_ranges[i]]["secondary"])

scoreboard(sheet, songs, color_ranges)

# column 6 for black border
unicolor_column2(sheet, column=6, start=0, end=len(songs)+1, width=3, color="000000")




# next... nominations, hell
# I'll definitely have to refactor this later
def nominations(sheet, songs, users):
    # starting from G column (7th column)
    col = 7

    counter = len(songs)

    color = colors_guide["yellow"]["secondary"]
    noms_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")


    for i in range(len(general_colors)):
        # important columns
        first = col + 3*i
        second = col + 3*i + 1
        last = col + 3*i + 2

        # important rows
        after_noms_row = nominations_num + 2

        # important colors
        decision_color_prim = colors_guide[general_colors[i]]["primary"]
        decision_color_sec = colors_guide[general_colors[i]]["secondary"]



        # header
        sheet.cell(row=1, column=first).value = "Nominated Song:"
        sheet.cell(row=1, column=second).value = "Nominated Song Link:"

        # nominations
        for j in range(nominations_num):
            sheet.cell(row=j+2, column=first).value = "h"
            sheet.cell(row=j+2, column=first).fill = noms_fill
            sheet.cell(row=j+2, column=second).value = "h"
            sheet.cell(row=j+2, column=second).fill = noms_fill

        # gap row
        sheet.cell(row=after_noms_row, column=first).value = "EMPTY"
        sheet.cell(row=after_noms_row, column=second).value = "EMPTY"

        # decisions
        how_many_colors = len([color for color in color_ranges if color == general_colors[i]])
        for j in range(how_many_colors):
            new_row = after_noms_row + (len(users)+3)*j + 1

            free_row = decision_space(
                sheet, 
                new_row, 
                first, 
                counter, 
                decision_color_prim, 
                decision_color_sec, 
                users)
            
            counter -= 1

        # last column check for bronze silver gold
        if counter == 3:
            # bronze
            free_row = decision_space(
                sheet, 
                free_row, 
                first, 
                counter, 
                colors_guide["bronze"]["primary"], 
                colors_guide["bronze"]["secondary"], 
                users)
            
            counter -= 1

            # silver
            free_row = decision_space(
                sheet, 
                free_row, 
                first, 
                counter, 
                colors_guide["silver"]["primary"], 
                colors_guide["silver"]["secondary"], 
                users)
            
            counter -= 1

            # gold 
            gold_color = colors_guide["gold"]["primary"]
            gold_fill = PatternFill(start_color=gold_color, end_color=gold_color, fill_type="solid")

            sheet.cell(row=free_row, column=first).value = "1st Place: "
            sheet.cell(row=free_row, column=first).fill = gold_fill
            sheet.cell(row=free_row, column=second).fill = gold_fill

        # black border
        unicolor_column2(sheet, column=last, start=0, end=max_black, width=3, color="000000")


    # add bronze, silver and gold
    for i in range(3):
        pass



            




nominations(sheet, songs, users)


# Save the workbook with the changes
workbook.save("output.xlsx")